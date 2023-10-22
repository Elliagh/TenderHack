import openpyxl

from src.algorithm import types
from src.algorithm.algorithm import getType

path = "Логи.xlsx"
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active
row = sheet_obj.max_row

for i in range(1, 10):
    cell_obj = sheet_obj.cell(row=i, column=3)
    log_info = cell_obj.value.split(" ")
    result = gqetType(log_info)
    print(result)
