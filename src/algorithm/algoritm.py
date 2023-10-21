import openpyxl

from src.algorithm import types

path = "Логи.xlsx"

wb_obj = openpyxl.load_workbook(path)

sheet_obj = wb_obj.active

row = sheet_obj.max_row

def getType(log_info: str) -> str:
    for i in log_info:
        if i in types.list_of_log_types:
            return types.list_of_log_types.get(i)


for i in range(1, 10):
    cell_obj = sheet_obj.cell(row = i, column = 3)
    log_info = cell_obj.value.split(" ")
    result = getType(log_info)
    print(result)


